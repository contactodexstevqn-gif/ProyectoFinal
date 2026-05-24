import logging
from html import escape

from django.conf import settings
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import Group, User
from django.core.paginator import Paginator
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from backend.pagination import OPCIONES_POR_PAGINA, obtener_por_pagina, parametros_sin_pagina
from backend.permissions import (
    GRUPO_VENDEDOR,
    admin_required,
    asegurar_grupos_base,
    es_administrador,
    rol_usuario,
)
from .forms import VendedorForm, VendedorEditForm
from .services import BrevoEmailError, enviar_correo_brevo

logger = logging.getLogger(__name__)


def iniciarSesion(request):
    if request.method == 'POST':
        nombreUsuario = request.POST.get('username')
        contraseña = request.POST.get('password')

        usuario = authenticate(
            request, username=nombreUsuario, password=contraseña)

        if usuario is not None:
            login(request, usuario)
            return redirect('dashboard')
        else:
            return render(request, 'public/login.html', {
                'error': 'Cuenta no encontrada'
            })

    return render(request, 'public/login.html')


def cerrarSesion(request):
    logout(request)
    return redirect('login')


def recuperarContraseña(request):
    if request.user.is_authenticated:
        return redirect('dashboard')

    if request.method == 'POST':
        correo = request.POST.get('email', '').strip()

        if not correo:
            messages.error(request, 'Ingresa el correo asociado a la cuenta.')
            return render(request, 'public/recuperar_contrasena.html')

        usuarios = User.objects.filter(email__iexact=correo)
        existe_cuenta = usuarios.exists()

        usuarios_encontrados = (
            ', '.join(usuario.username for usuario in usuarios)
            if existe_cuenta
            else 'No registrado'
        )

        enlace_edicion = 'No disponible'
        if existe_cuenta:
            primer_usuario = usuarios.first()
            enlace_edicion = request.build_absolute_uri(
                reverse('editar_vendedor', args=[primer_usuario.id])
            )

        fecha_solicitud = timezone.localtime(
            timezone.now()
        ).strftime('%d/%m/%Y %I:%M %p')

        ip_cliente = request.META.get(
            'HTTP_X_FORWARDED_FOR',
            request.META.get('REMOTE_ADDR', 'No disponible')
        )
        if ',' in ip_cliente:
            ip_cliente = ip_cliente.split(',')[0].strip()

        asunto = 'Solicitud de recuperacion de contraseña'
        mensaje = (
            'Se recibio una solicitud de recuperacion de contraseña desde el sistema.\n\n'
            f'Correo ingresado: {correo}\n'
            f'Cuenta registrada: {"Si" if existe_cuenta else "No"}\n'
            f'Usuario(s): {usuarios_encontrados}\n'
            f'Enlace para cambiar contraseña: {enlace_edicion}\n'
            f'Fecha y hora: {fecha_solicitud}\n'
            f'IP aproximada: {ip_cliente}\n\n'
            'El usuario no recibio ningun correo automaticamente. '
            'El administrador debe validar la solicitud y, si corresponde, cambiar la contraseña '
            'desde Vendedores > Editar vendedor.'
        )
        mensaje_html = (
            '<h2>Solicitud de recuperacion de contrase&ntilde;a</h2>'
            '<p>Se recibio una solicitud de recuperacion de contrase&ntilde;a '
            'desde el sistema.</p>'
            '<ul>'
            f'<li><strong>Correo ingresado:</strong> {escape(correo)}</li>'
            f'<li><strong>Cuenta registrada:</strong> {"Si" if existe_cuenta else "No"}</li>'
            f'<li><strong>Usuario(s):</strong> {escape(usuarios_encontrados)}</li>'
            f'<li><strong>Fecha y hora:</strong> {escape(fecha_solicitud)}</li>'
            f'<li><strong>IP aproximada:</strong> {escape(ip_cliente)}</li>'
            '</ul>'
            '<p><strong>Enlace para cambiar contrase&ntilde;a:</strong><br>'
            f'<a href="{escape(enlace_edicion)}">{escape(enlace_edicion)}</a></p>'
            '<p>El usuario no recibio ningun correo automaticamente. '
            'El administrador debe validar la solicitud y, si corresponde, '
            'cambiar la contrase&ntilde;a desde Vendedores &gt; Editar vendedor.</p>'
        )

        correo_admin = getattr(settings, 'ADMIN_RECOVERY_EMAIL', '')

        if not correo_admin:
            logger.error(
                'No se pudo enviar recuperacion: falta ADMIN_RECOVERY_EMAIL.'
            )
        else:
            try:
                enviar_correo_brevo(
                    correo_admin,
                    'Administrador',
                    asunto,
                    mensaje_html,
                    mensaje,
                )
            except BrevoEmailError as error:
                logger.exception(
                    'Error enviando correo de recuperacion con Brevo: %s',
                    error
                )

        messages.info(
            request,
            'Si el correo esta registrado, el administrador recibira tu solicitud.'
        )
        return redirect('recuperar_contrasena')

    return render(request, 'public/recuperar_contrasena.html')


recuperarContrasena = recuperarContraseña

@login_required
@admin_required
def gestionUsuarios(request):
    usuarios = User.objects.filter(is_superuser=False).order_by('-date_joined')

    total_usuarios = usuarios.count()
    usuarios_activos = usuarios.filter(is_active=True).count()
    usuarios_inactivos = usuarios.filter(is_active=False).count()

    per_page, per_page_int = obtener_por_pagina(request)
    paginator = Paginator(usuarios, per_page_int)
    pagina = request.GET.get('page')
    usuarios_pagina = paginator.get_page(pagina)
    query_params = parametros_sin_pagina(request, ['page'])

    return render(request, 'usuarios/usuarios.html', {
        'usuarios': usuarios_pagina,
        'total_usuarios': total_usuarios,
        'usuarios_activos': usuarios_activos,
        'usuarios_inactivos': usuarios_inactivos,
        'es_admin': es_administrador(request.user),
        'rol_usuario': rol_usuario(request.user),
        'query_params': query_params,
        'per_page': per_page,
        'per_page_options': OPCIONES_POR_PAGINA,
    })


@login_required
@admin_required
def crearVendedor(request):
    asegurar_grupos_base()

    if request.method == 'POST':
        form = VendedorForm(request.POST)

        if form.is_valid():
            vendedor = form.save(commit=False)

            contra = form.cleaned_data.get('contra')
            vendedor.set_password(contra)

            vendedor.is_superuser = False
            vendedor.is_staff = False
            vendedor.is_active = True

            vendedor.save()

            grupo_vendedor = Group.objects.get(name=GRUPO_VENDEDOR)
            vendedor.groups.add(grupo_vendedor)

            messages.success(request, 'Usuario creado correctamente.')
            return redirect('usuarios')
    else:
        form = VendedorForm()

    return render(request, 'usuarios/crear_vendedor.html', {
        'form': form,
        'es_admin': es_administrador(request.user),
        'rol_usuario': rol_usuario(request.user),
    })


@login_required
@admin_required
def editarVendedor(request, usuario_id):
    vendedor = get_object_or_404(User, id=usuario_id, is_superuser=False)

    if request.method == 'POST':
        form = VendedorEditForm(request.POST, instance=vendedor)

        if form.is_valid():
            cambio_contrasena = bool(form.cleaned_data.get('nueva_contrasena'))
            form.save()
            if cambio_contrasena:
                messages.success(
                    request, f'El vendedor {vendedor.username} fue actualizado y su contraseña fue cambiada correctamente.')
            else:
                messages.success(
                    request, f'El vendedor {vendedor.username} fue actualizado correctamente.')
            return redirect('usuarios')

        messages.error(
            request, 'No se pudo actualizar el vendedor. Revisa los datos ingresados.')
    else:
        form = VendedorEditForm(instance=vendedor)

    return render(request, 'usuarios/editar_vendedor.html', {
        'form': form,
        'vendedor': vendedor,
        'es_admin': es_administrador(request.user),
        'rol_usuario': rol_usuario(request.user),
    })


@login_required
@admin_required
def exportarVendedoresExcel(request):
    usuarios = User.objects.filter(is_superuser=False).order_by('-date_joined')

    total_usuarios = usuarios.count()
    usuarios_activos = usuarios.filter(is_active=True).count()
    usuarios_inactivos = usuarios.filter(is_active=False).count()

    wb = Workbook()
    ws = wb.active
    ws.title = 'Vendedores'

    encabezados = [
        'Nombre',
        'Usuario',
        'Correo',
        'Estado',
        'Fecha de registro',
        'Ultimo acceso'
    ]

    ws.append(['Listado de vendedores'])
    ws.append([
        f'Total: {total_usuarios}',
        f'Activos: {usuarios_activos}',
        f'Inactivos: {usuarios_inactivos}',
        f'Generado: {timezone.localtime(timezone.now()).strftime("%d/%m/%Y %I:%M %p")}'
    ])
    ws.append([])
    ws.append(encabezados)

    for usuario in usuarios:
        nombre = usuario.get_full_name() or 'Sin nombre'
        ultimo_acceso = (
            timezone.localtime(usuario.last_login).strftime(
                '%d/%m/%Y %I:%M %p')
            if usuario.last_login else 'Sin acceso'
        )

        ws.append([
            nombre,
            usuario.username,
            usuario.email or 'Sin correo',
            'Activo' if usuario.is_active else 'Inactivo',
            timezone.localtime(usuario.date_joined).strftime(
                '%d/%m/%Y %I:%M %p'),
            ultimo_acceso,
        ])

    titulo_fill = PatternFill('solid', fgColor='D41473')
    encabezado_fill = PatternFill('solid', fgColor='FCE7F3')
    resumen_fill = PatternFill('solid', fgColor='FFF1F8')
    borde_color = Side(style='thin', color='E5E7EB')

    titulo_font = Font(color='FFFFFF', bold=True, size=15)
    encabezado_font = Font(color='111827', bold=True)
    resumen_font = Font(color='4B5563', bold=True)
    texto_font = Font(color='111827')

    center = Alignment(horizontal='center', vertical='center')
    left = Alignment(horizontal='left', vertical='center')

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
    ws['A1'].fill = titulo_fill
    ws['A1'].font = titulo_font
    ws['A1'].alignment = center
    ws.row_dimensions[1].height = 28

    for cell in ws[2]:
        cell.fill = resumen_fill
        cell.font = resumen_font
        cell.alignment = left
        cell.border = Border(left=borde_color, right=borde_color,
                             top=borde_color, bottom=borde_color)

    for cell in ws[4]:
        cell.fill = encabezado_fill
        cell.font = encabezado_font
        cell.alignment = center
        cell.border = Border(left=borde_color, right=borde_color,
                             top=borde_color, bottom=borde_color)

    for row in ws.iter_rows(min_row=5):
        for cell in row:
            cell.font = texto_font
            cell.alignment = left
            cell.border = Border(
                left=borde_color, right=borde_color, top=borde_color, bottom=borde_color)

    anchos = {
        'A': 30,
        'B': 22,
        'C': 32,
        'D': 14,
        'E': 24,
        'F': 24,
    }

    for columna, ancho in anchos.items():
        ws.column_dimensions[columna].width = ancho

    ws.freeze_panes = 'A5'
    ws.auto_filter.ref = f'A4:F{ws.max_row}'

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="vendedores.xlsx"'

    wb.save(response)
    return response


@login_required
@admin_required
def cambiarEstadoUsuario(request, usuario_id):
    if request.method != 'POST':
        return redirect('usuarios')

    usuario = get_object_or_404(User, id=usuario_id, is_superuser=False)

    if usuario == request.user:
        messages.error(request, 'No puedes desactivar tu propio usuario.')
        return redirect('usuarios')

    usuario.is_active = not usuario.is_active
    usuario.save(update_fields=['is_active'])

    if usuario.is_active:
        messages.success(
            request, f'El usuario {usuario.username} fue activado correctamente.')
    else:
        messages.success(
            request, f'El usuario {usuario.username} fue desactivado correctamente.')

    return redirect('usuarios')
