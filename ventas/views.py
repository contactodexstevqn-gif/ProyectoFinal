from datetime import datetime, time, timedelta
from decimal import InvalidOperation

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.core.exceptions import ValidationError
from django.core.paginator import Paginator
from django.db import transaction
from django.db.models import Count, Max, Q, Sum
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.utils.dateparse import parse_date
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from backend.permissions import GRUPO_ADMINISTRADOR, admin_required, es_administrador, rol_usuario
from productos.models import Producto
from .models import Cliente, Venta


def rango_dia(fecha):
    zona = timezone.get_current_timezone()
    inicio = datetime.combine(fecha, time.min)
    fin = inicio + timedelta(days=1)

    if timezone.is_naive(inicio):
        inicio = timezone.make_aware(inicio, zona)

    if timezone.is_naive(fin):
        fin = timezone.make_aware(fin, zona)

    return inicio, fin


def rango_mes(fecha):
    zona = timezone.get_current_timezone()
    inicio = datetime(fecha.year, fecha.month, 1)

    if fecha.month == 12:
        fin = datetime(fecha.year + 1, 1, 1)
    else:
        fin = datetime(fecha.year, fecha.month + 1, 1)

    if timezone.is_naive(inicio):
        inicio = timezone.make_aware(inicio, zona)

    if timezone.is_naive(fin):
        fin = timezone.make_aware(fin, zona)

    return inicio, fin


def vendedores_registrados():
    return User.objects.filter(
        is_active=True
    ).exclude(
        is_superuser=True
    ).exclude(
        groups__name=GRUPO_ADMINISTRADOR
    ).distinct().order_by('first_name', 'last_name', 'username')


def ventas_permitidas(usuario, es_admin):
    ventas = Venta.objects.select_related(
        'producto',
        'producto__categoria',
        'vendedor',
        'cliente'
    )

    if not es_admin:
        ventas = ventas.filter(vendedor=usuario)

    return ventas


def aplicar_filtros_ventas(request, ventas, es_admin):
    query = request.GET.get('q', '').strip()
    fecha = request.GET.get('fecha', '').strip()
    vendedor_id = request.GET.get('vendedor', '').strip()

    if query:
        ventas = ventas.filter(
            Q(producto__nombre__icontains=query) |
            Q(producto__categoria__nombre__icontains=query) |
            Q(producto__color__icontains=query) |
            Q(producto__talla__icontains=query) |
            Q(vendedor__username__icontains=query) |
            Q(vendedor__first_name__icontains=query) |
            Q(vendedor__last_name__icontains=query) |
            Q(cliente__documento__icontains=query) |
            Q(cliente__nombre__icontains=query) |
            Q(cliente__correo__icontains=query)
        )

    if fecha:
        fecha_filtro = parse_date(fecha)

        if fecha_filtro:
            inicio_fecha, fin_fecha = rango_dia(fecha_filtro)

            ventas = ventas.filter(
                fecha__gte=inicio_fecha,
                fecha__lt=fin_fecha
            )

    if es_admin and vendedor_id:
        ventas = ventas.filter(vendedor_id=vendedor_id)

    return ventas, query, fecha, vendedor_id


def obtener_cliente_venta(request):
    cliente_modo = request.POST.get('cliente_modo', 'rapida').strip()
    cliente_id = request.POST.get('cliente_id', '').strip()
    documento_cliente = request.POST.get('documento_cliente', '').strip()
    nombre_cliente = request.POST.get('nombre_cliente', '').strip()
    correo_cliente = request.POST.get('correo_cliente', '').strip()
    telefono_cliente = request.POST.get('telefono_cliente', '').strip()

    if cliente_modo == 'rapida':
        return None

    if cliente_modo == 'existente':
        if not cliente_id:
            raise ValidationError('Debes seleccionar un cliente existente.')

        cliente = Cliente.objects.filter(id=cliente_id).first()

        if cliente is None:
            raise ValidationError('El cliente seleccionado no existe.')

        return cliente

    if cliente_modo == 'nuevo':
        if not documento_cliente:
            raise ValidationError('Debes ingresar el documento del cliente.')

        if not nombre_cliente:
            raise ValidationError('Debes ingresar el nombre del cliente.')

        cliente, creado = Cliente.objects.get_or_create(
            documento=documento_cliente,
            defaults={
                'nombre': nombre_cliente,
                'correo': correo_cliente or None,
                'telefono': telefono_cliente or None
            }
        )

        if not creado:
            cliente.nombre = nombre_cliente
            cliente.correo = correo_cliente or None
            cliente.telefono = telefono_cliente or None
            cliente.save(update_fields=['nombre', 'correo', 'telefono'])

        return cliente

    raise ValidationError('La opción de cliente seleccionada no es válida.')


@login_required(login_url='login')
def nueva_venta(request):
    es_admin = es_administrador(request.user)

    productos = Producto.objects.select_related('categoria').all().order_by('nombre')
    productos_disponibles = productos.filter(stock__gt=0)
    vendedores = vendedores_registrados()
    clientes = Cliente.objects.all().order_by('nombre')

    error = None
    vendedor_id_seleccionado = ''
    cliente_modo_seleccionado = 'rapida'
    cliente_id_seleccionado = ''
    documento_cliente = ''
    nombre_cliente = ''
    correo_cliente = ''
    telefono_cliente = ''

    if request.method == 'POST':
        vendedor_id = request.POST.get('vendedor', '').strip()
        producto_ids = request.POST.getlist('producto')
        cantidades_raw = request.POST.getlist('cantidad')
        cliente_modo_seleccionado = request.POST.get('cliente_modo', 'rapida').strip()
        cliente_id_seleccionado = request.POST.get('cliente_id', '').strip()
        documento_cliente = request.POST.get('documento_cliente', '').strip()
        nombre_cliente = request.POST.get('nombre_cliente', '').strip()
        correo_cliente = request.POST.get('correo_cliente', '').strip()
        telefono_cliente = request.POST.get('telefono_cliente', '').strip()

        vendedor_id_seleccionado = vendedor_id
        vendedor = request.user
        cliente = None

        if es_admin:
            if not vendedor_id:
                error = 'Debes seleccionar el vendedor que realizó la venta.'
            else:
                vendedor = vendedores.filter(id=vendedor_id).first()

                if vendedor is None:
                    error = 'El vendedor seleccionado no es válido.'

        if not error:
            try:
                cliente = obtener_cliente_venta(request)
            except ValidationError as e:
                error = e.messages[0] if hasattr(e, 'messages') and e.messages else 'No se pudo validar el cliente.'

        productos_agrupados = {}

        if not error:
            for index, producto_id in enumerate(producto_ids):
                producto_id = str(producto_id).strip()
                cantidad_raw = cantidades_raw[index] if index < len(cantidades_raw) else ''

                if not producto_id and not cantidad_raw:
                    continue

                if not producto_id:
                    error = 'Debes seleccionar un producto en todas las filas.'
                    break

                try:
                    producto_id_int = int(producto_id)
                    cantidad = int(cantidad_raw)
                except (TypeError, ValueError):
                    error = 'Todas las cantidades deben ser números válidos.'
                    break

                if cantidad <= 0:
                    error = 'Todas las cantidades deben ser mayores a 0.'
                    break

                productos_agrupados[producto_id_int] = productos_agrupados.get(producto_id_int, 0) + cantidad

            if not error and not productos_agrupados:
                error = 'Debes agregar al menos un producto a la venta.'

        if not error:
            try:
                with transaction.atomic():
                    productos_en_venta = Producto.objects.select_for_update().filter(
                        id__in=productos_agrupados.keys()
                    )

                    productos_map = {
                        producto.id: producto
                        for producto in productos_en_venta
                    }

                    for producto_id, cantidad in productos_agrupados.items():
                        producto = productos_map.get(producto_id)

                        if producto is None:
                            raise ValidationError('Uno de los productos seleccionados no existe.')

                        if cantidad > producto.stock:
                            raise ValidationError(
                                f'No hay suficiente stock para {producto.nombre}. Stock disponible: {producto.stock}.'
                            )

                    for producto_id, cantidad in productos_agrupados.items():
                        Venta.objects.create(
                            producto=productos_map[producto_id],
                            vendedor=vendedor,
                            cliente=cliente,
                            cantidad=cantidad
                        )

                messages.success(request, 'Venta registrada correctamente.')
                return redirect('nueva_venta')

            except ValidationError as e:
                error = e.messages[0] if hasattr(e, 'messages') and e.messages else 'No se pudo registrar la venta.'
            except (ValueError, InvalidOperation):
                error = 'No se pudo registrar la venta. Revisa los datos e intenta de nuevo.'

    ventas_base = ventas_permitidas(request.user, es_admin)
    ventas_recientes = ventas_base.order_by('-fecha')[:5]

    hoy = timezone.localdate()
    inicio_hoy, fin_hoy = rango_dia(hoy)
    inicio_mes, fin_mes = rango_mes(hoy)

    ventas_hoy = ventas_base.filter(fecha__gte=inicio_hoy, fecha__lt=fin_hoy)
    ventas_mes = ventas_base.filter(fecha__gte=inicio_mes, fecha__lt=fin_mes)

    total_ventas = ventas_base.aggregate(total=Sum('total'))['total'] or 0
    total_hoy = ventas_hoy.aggregate(total=Sum('total'))['total'] or 0
    total_mes = ventas_mes.aggregate(total=Sum('total'))['total'] or 0
    unidades_vendidas = ventas_base.aggregate(total=Sum('cantidad'))['total'] or 0

    return render(request, 'nueva_venta.html', {
        'productos': productos,
        'productos_disponibles': productos_disponibles,
        'vendedores_registrados': vendedores,
        'clientes_registrados': clientes,
        'ventas_recientes': ventas_recientes,
        'error': error,
        'es_admin': es_admin,
        'rol_usuario': rol_usuario(request.user),
        'total_ventas': total_ventas,
        'total_hoy': total_hoy,
        'total_mes': total_mes,
        'unidades_vendidas': unidades_vendidas,
        'vendedor_id_seleccionado': vendedor_id_seleccionado,
        'cliente_modo_seleccionado': cliente_modo_seleccionado,
        'cliente_id_seleccionado': cliente_id_seleccionado,
        'documento_cliente': documento_cliente,
        'nombre_cliente': nombre_cliente,
        'correo_cliente': correo_cliente,
        'telefono_cliente': telefono_cliente,
    })


@login_required(login_url='login')
def historial_ventas(request):
    es_admin = es_administrador(request.user)

    ventas_base = ventas_permitidas(request.user, es_admin)

    ventas_filtradas, query, fecha, vendedor_id = aplicar_filtros_ventas(
        request,
        ventas_base,
        es_admin
    )

    total_filtrado = ventas_filtradas.aggregate(total=Sum('total'))['total'] or 0
    unidades_filtradas = ventas_filtradas.aggregate(total=Sum('cantidad'))['total'] or 0
    cantidad_ventas = ventas_filtradas.count()

    ventas = ventas_filtradas.order_by('-fecha')
    vendedores = vendedores_registrados() if es_admin else User.objects.none()

    return render(request, 'historial_ventas.html', {
        'ventas': ventas,
        'vendedores': vendedores,
        'query': query,
        'fecha': fecha,
        'vendedor_id': vendedor_id,
        'es_admin': es_admin,
        'rol_usuario': rol_usuario(request.user),
        'total_filtrado': total_filtrado,
        'unidades_filtradas': unidades_filtradas,
        'cantidad_ventas': cantidad_ventas,
    })


@login_required(login_url='login')
@admin_required
def clientes(request):
    query = request.GET.get('q', '').strip()

    clientes_lista = Cliente.objects.annotate(
        cantidad_ventas=Count('ventas'),
        total_comprado=Sum('ventas__total'),
        ultima_compra=Max('ventas__fecha')
    ).order_by('nombre')

    if query:
        clientes_lista = clientes_lista.filter(
            Q(documento__icontains=query) |
            Q(nombre__icontains=query) |
            Q(correo__icontains=query) |
            Q(telefono__icontains=query)
        )

    total_clientes = clientes_lista.count()
    clientes_con_compras = clientes_lista.filter(cantidad_ventas__gt=0).count()
    total_compras = clientes_lista.aggregate(total=Sum('ventas__total'))['total'] or 0

    paginator = Paginator(clientes_lista, 15)
    pagina = request.GET.get('page')
    clientes_pagina = paginator.get_page(pagina)

    return render(request, 'clientes.html', {
        'clientes': clientes_pagina,
        'query': query,
        'total_clientes': total_clientes,
        'clientes_con_compras': clientes_con_compras,
        'total_compras': total_compras,
        'es_admin': es_administrador(request.user),
        'rol_usuario': rol_usuario(request.user),
    })


@login_required(login_url='login')
@admin_required
def detalle_cliente(request, cliente_id):
    cliente = get_object_or_404(Cliente, id=cliente_id)

    ventas_cliente = Venta.objects.select_related(
        'producto',
        'producto__categoria',
        'vendedor',
        'cliente'
    ).filter(
        cliente=cliente
    ).order_by('-fecha')

    total_comprado = ventas_cliente.aggregate(total=Sum('total'))['total'] or 0
    cantidad_ventas = ventas_cliente.count()
    unidades_compradas = ventas_cliente.aggregate(total=Sum('cantidad'))['total'] or 0
    productos_diferentes = ventas_cliente.values('producto_id').distinct().count()

    paginator = Paginator(ventas_cliente, 15)
    pagina = request.GET.get('page')
    ventas_pagina = paginator.get_page(pagina)

    return render(request, 'detalle_cliente.html', {
        'cliente': cliente,
        'ventas': ventas_pagina,
        'total_comprado': total_comprado,
        'cantidad_ventas': cantidad_ventas,
        'unidades_compradas': unidades_compradas,
        'productos_diferentes': productos_diferentes,
        'es_admin': es_administrador(request.user),
        'rol_usuario': rol_usuario(request.user),
    })

@login_required(login_url='login')
@admin_required
def editar_cliente(request, cliente_id):
    cliente = get_object_or_404(Cliente, id=cliente_id)

    if request.method == 'POST':
        documento = request.POST.get('documento', '').strip()
        nombre = request.POST.get('nombre', '').strip()
        correo = request.POST.get('correo', '').strip()
        telefono = request.POST.get('telefono', '').strip()

        if not documento or not nombre:
            messages.error(request, 'Documento y nombre son obligatorios.')
            return redirect('editar_cliente', cliente_id=cliente.id)

        cliente.documento = documento
        cliente.nombre = nombre
        cliente.correo = correo or None
        cliente.telefono = telefono or None
        cliente.save(update_fields=['documento', 'nombre', 'correo', 'telefono'])

        messages.success(request, 'Cliente actualizado correctamente.')
        return redirect('clientes')

    return render(request, 'editar_cliente.html', {
        'cliente': cliente,
        'es_admin': es_administrador(request.user),
        'rol_usuario': rol_usuario(request.user),
    })


@login_required(login_url='login')
@admin_required
def cambiar_estado_cliente(request, cliente_id):
    cliente = get_object_or_404(Cliente, id=cliente_id)

    if request.method == 'POST':
        cliente.activo = not cliente.activo
        cliente.save(update_fields=['activo'])

        if cliente.activo:
            messages.success(request, 'Cliente activado correctamente.')
        else:
            messages.success(request, 'Cliente desactivado correctamente.')

    return redirect('clientes')


@login_required(login_url='login')
def exportar_ventas(request):
    es_admin = es_administrador(request.user)

    ventas_base = ventas_permitidas(request.user, es_admin)

    ventas_filtradas, query, fecha, vendedor_id = aplicar_filtros_ventas(
        request,
        ventas_base,
        es_admin
    )

    ventas = ventas_filtradas.order_by('-fecha')

    nombre_archivo = 'historial_ventas.xlsx' if es_admin else 'mis_ventas.xlsx'

    total_filtrado = ventas.aggregate(total=Sum('total'))['total'] or 0
    unidades_filtradas = ventas.aggregate(total=Sum('cantidad'))['total'] or 0
    cantidad_ventas = ventas.count()

    nombre_vendedor = 'Todos'

    if es_admin and vendedor_id:
        vendedor = User.objects.filter(id=vendedor_id).first()

        if vendedor:
            nombre_vendedor = vendedor.get_full_name() or vendedor.username

    if not es_admin:
        nombre_vendedor = request.user.get_full_name() or request.user.username

    wb = Workbook()
    ws = wb.active
    ws.title = 'Historial ventas'

    encabezados = [
        'Fecha',
        'Producto',
        'Categoria',
        'Color',
        'Talla',
        'Cantidad',
        'Total',
        'Vendedor',
        'Documento cliente',
        'Cliente',
        'Correo cliente'
    ]

    ws.append(['Historial de ventas'])
    ws.append([
        f"Busqueda: {query or 'Todos'}",
        f"Fecha: {fecha or 'Todas'}",
        f"Vendedor: {nombre_vendedor}",
        f"Ventas: {cantidad_ventas}",
        f"Unidades: {unidades_filtradas}",
        f"Total: ${int(total_filtrado):,}".replace(',', '.')
    ])
    ws.append([])
    ws.append(encabezados)

    for venta in ventas:
        ws.append([
            timezone.localtime(venta.fecha).strftime('%d/%m/%Y %I:%M %p'),
            venta.producto.nombre,
            venta.producto.categoria.nombre if venta.producto.categoria else 'Sin categoria',
            venta.producto.color,
            venta.producto.talla,
            venta.cantidad,
            float(venta.total or 0),
            venta.vendedor.username if venta.vendedor else '',
            venta.cliente.documento if venta.cliente else '',
            venta.cliente.nombre if venta.cliente else 'Sin cliente',
            venta.cliente.correo if venta.cliente and venta.cliente.correo else '',
        ])

    titulo_fill = PatternFill('solid', fgColor='D41473')
    encabezado_fill = PatternFill('solid', fgColor='FCE7F3')
    resumen_fill = PatternFill('solid', fgColor='FFF1F8')
    borde_color = Side(style='thin', color='E5E7EB')

    titulo_font = Font(color='FFFFFF', bold=True, size=15)
    encabezado_font = Font(color='111827', bold=True)
    resumen_font = Font(color='4B5563', bold=True)
    texto_font = Font(color='111827')

    center = Alignment(horizontal='center', vertical='center')
    left = Alignment(horizontal='left', vertical='center')

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=11)
    ws['A1'].fill = titulo_fill
    ws['A1'].font = titulo_font
    ws['A1'].alignment = center
    ws.row_dimensions[1].height = 28

    for cell in ws[2]:
        cell.fill = resumen_fill
        cell.font = resumen_font
        cell.alignment = left
        cell.border = Border(
            left=borde_color,
            right=borde_color,
            top=borde_color,
            bottom=borde_color
        )

    for cell in ws[4]:
        cell.fill = encabezado_fill
        cell.font = encabezado_font
        cell.alignment = center
        cell.border = Border(
            left=borde_color,
            right=borde_color,
            top=borde_color,
            bottom=borde_color
        )

    for row in ws.iter_rows(min_row=5):
        for cell in row:
            cell.font = texto_font
            cell.alignment = left
            cell.border = Border(
                left=borde_color,
                right=borde_color,
                top=borde_color,
                bottom=borde_color
            )

    for row in ws.iter_rows(min_row=5, min_col=6, max_col=7):
        for cell in row:
            cell.alignment = center

    for cell in ws['G'][4:]:
        cell.number_format = '"$"#,##0'

    anchos = {
        'A': 23,
        'B': 28,
        'C': 22,
        'D': 16,
        'E': 12,
        'F': 12,
        'G': 16,
        'H': 22,
        'I': 20,
        'J': 28,
        'K': 30,
    }

    for columna, ancho in anchos.items():
        ws.column_dimensions[columna].width = ancho

    ws.freeze_panes = 'A5'
    ws.auto_filter.ref = f'A4:K{ws.max_row}'

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

    response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'

    wb.save(response)

    return response